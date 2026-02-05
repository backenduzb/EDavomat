import openpyxl as pyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

from adminpage.models import Statistics
from students.models import Students


def write_excel(stat_id: int, file_path: str | None = None):
    
    stat = (
        Statistics.objects
        .select_related("school")
        .get(id=stat_id)
    )

    school_name = stat.school.name if stat.school else "Maktab"
    day_str = stat.created_at.isoformat()

    reason_qs = (
        stat.reason_students
        .select_related("_class__name")
        .all()
    )
    no_reason_qs = (
        stat.no_reason_students
        .select_related("_class__name")
        .all()
    )

    rows = []

    for s in reason_qs:
        rows.append((
            s.full_name,
            (s._class.name.name if s._class and s._class.name else "-"),
            "Sababli",
            (s.sababi or "-"),
        ))

    for s in no_reason_qs:
        rows.append((
            s.full_name,
            (s._class.name.name if s._class and s._class.name else "-"),
            "Sababsiz",
            (s.sababi or "-"),
        ))

    wb = pyxl.Workbook()
    ws = wb.active
    ws.title = "Kelmaganlar"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    body_font = Font(size=11, color="111111")

    title_fill = PatternFill("solid", fgColor="0F172A")   
    header_fill = PatternFill("solid", fgColor="16A34A")  
    alt_fill = PatternFill("solid", fgColor="F1F5F9")     
    status_reason_fill = PatternFill("solid", fgColor="DCFCE7")   
    status_noreason_fill = PatternFill("solid", fgColor="FEE2E2") 

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = ["№", "F.I.Sh", "Sinfi", "Holati", "Sababi", "Sana"]
    col_widths = [6, 34, 12, 12, 45, 12]

    title_text = f"{school_name} | Kelmagan o'quvchilar ro'yxati | {day_str}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 34

    header_row = 3
    for idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = col_widths[idx - 1]
    ws.row_dimensions[header_row].height = 22

    start_row = header_row + 1
    for i, (full_name, class_name, status, reason_text) in enumerate(rows, start=1):
        r = start_row + (i - 1)

        values = [i, full_name, class_name, status, reason_text, day_str]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = body_font
            cell.border = border

            if col in (1, 3, 4, 6):
                cell.alignment = center
            else:
                cell.alignment = left

            if i % 2 == 0:
                cell.fill = alt_fill

            if col == 4:
                if status == "Sababli":
                    cell.fill = status_reason_fill
                else:
                    cell.fill = status_noreason_fill

        ws.row_dimensions[r].height = 20

    if not rows:
        r = start_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns))
        cell = ws.cell(row=r, column=1, value="Bu sana uchun kelmaganlar topilmadi.")
        cell.alignment = center
        cell.font = Font(size=12, bold=True, color="334155")
        ws.row_dimensions[r].height = 24

    ws.freeze_panes = ws["A4"]

    ws.auto_filter.ref = f"A{header_row}:F{max(start_row, start_row + len(rows) - 1)}"

    if file_path is None:
        file_path = f"kelmaganlar_{school_name}_{day_str}.xlsx".replace(" ", "_").replace("/", "-")

    wb.save(file_path)
    return file_path, len(rows)
